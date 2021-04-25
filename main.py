from openpyxl import load_workbook
from openpyxl.writer.excel import save_workbook
import pathlib


def is_number_norm(string):
    if string:
        return string.isdigit()
    else:
        return False


def is_number(string):
    if string:
        string = string[0:-1]
        return string.isdigit()
    else:
        return False


def max_row(wsf):
    iterator = 1
    string = 0
    while (iterator > 0) and (iterator - string < 30):
        # print('i**** - ' + i.__str__())
        # print(string)
        if is_number(wsf['A' + iterator.__str__()].value):
            # print('i - ' + i.__str__())
            string = iterator
        iterator += 1
    return iterator


def cleaner03(file_name):

    FILE_NAME_IN = file_name + '.xlsx'
    FILE_NAME_OUT = file_name + ' исправленный.xlsx'
    FILE_NAME_OUT_LIST = file_name + ' исправленный разбит на листы.xlsx'

    wbf1 = load_workbook(FILE_NAME_IN)
    wsf1 = wbf1.active

    # let_name = ['B', 'C', 'D', 'E']
    C_FJ_letters = ['C', 'F', 'G', 'H', 'I', 'J']
    FJ_letters = ['F', 'G', 'H', 'I', 'J']
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    str_f1 = max_row(wsf1)  # количество строк
    i = 1
    while i < str_f1:
        wsf1['K' + i.__str__()].value = ''
        for j in range(i + 1, str_f1):
            if wsf1['A' + j.__str__()].value and is_number(wsf1['A' + j.__str__()].value) or j == str_f1 - 1:  # пос-
                # леднее условие для того чтобы последнюю строку чистило
                if (j - i) > 1:
                    for k in range(i + 1, j):
                        # Очистка столбца C, F-J (Завершено)
                        # нужно условие, чтобы не поднимались 1.ххх 2.ххх ... 5.ххх на верх
                        # ниже доперемещал, сплитил по ;
                        if wsf1['F' + k.__str__()].value.__str__().split(' ')[0] == '1.':  # 6040
                            # работаем с областью описания, склеиваем все в одну ячейку и
                            # очищаем ячейки откуда переместили
                            for temp in range(k, j):
                                for let in FJ_letters:
                                    # print(wsf1[let + temp.__str__()].value)
                                    if wsf1[let + temp.__str__()].value and wsf1['F' + k.__str__()].value:
                                        if not (let == 'F' and temp == k):
                                            wsf1['F' + k.__str__()].value = wsf1['F' + k.__str__()].value.__str__() + \
                                                                            wsf1[let + temp.__str__()].value.__str__()
                                            wsf1[let + temp.__str__()].value = None
                            # работаем с wsf1['F' + k.__str__()].value
                            # print(wsf1['F' + k.__str__()].value)
                            break
                        if wsf1['B' + k.__str__()].value is None:
                            for fj_let in C_FJ_letters:
                                if wsf1[fj_let + k.__str__()].value is not None and \
                                        wsf1[fj_let + k.__str__()].value is not None and \
                                        wsf1[fj_let + i.__str__()].value is not None:
                                    wsf1[fj_let + i.__str__()].value = wsf1[fj_let + i.__str__()].value + \
                                                                       wsf1[fj_let + k.__str__()].value
                                    wsf1[fj_let + k.__str__()].value = None
                i = j - 1
                break
        i += 1

    # start склеиваем и двигаем названия по типу "серия" влево (по образцу)
    i = 1
    while i < str_f1:
        if not (wsf1['A' + i.__str__()].value != ' ' and
                wsf1['A' + i.__str__()].value) and \
                wsf1['B' + i.__str__()].value:
            wsf1['A' + i.__str__()].value = ''
            for let in letters:
                if wsf1[let + i.__str__()].value:
                    # print(1)
                    wsf1['A' + i.__str__()].value = wsf1['A' + i.__str__()].value + wsf1[let + i.__str__()].value
                    wsf1[let + i.__str__()].value = None
        i += 1
    # end склеиваем и двигаем названия по типу "серия" влево (по образцу)

    # start разделяем и двигаем Штуки из столбцов F-J и поднимаем их на строку выше
    i = 1
    while i < str_f1:
        if wsf1['F' + i.__str__()].value.__str__().split(' ')[0] == '1.':
            temp = wsf1['F' + i.__str__()].value.__str__().split(';')  # мы дробим на пять штук
            # и записуем их в массив
            # print(i)
            # print(temp)
            for j in range(len(temp)):  # не везде 5 параметров
                # print(temp[j])
                # возможно ниже стоит добавлять в строки ';' ДЛя Лехи
                # print(i)
                # print(temp[j])
                wsf1[FJ_letters[j] + (i - 1).__str__()].value = temp[j]  # каждую штуку мы будем записывать
                # по своей колонке
            wsf1['F' + i.__str__()].value = None  # обнуление
        i += 1
    # end разделяем и двигаем Штуки из столбцов F-J и поднимаем их на строку выше

    # start доочистка
    # если есть "ТИПА", то (это N-типа) переместить в  ячейку B(i-1) и удалить
    # если нет "ТИПА", то (это более 30А) переместить это в ячейку выше (очистить откуда) и потом
    # ячейки F-J поднять выше

    # FILE_NAME_OUT = 'f1_edited_test.xlsx'
    # wb = load_workbook(FILE_NAME_IN)
    wb = wbf1
    ws = wb.active

    # print(ws.cell(row=3847, column=1).value)

    # str_f1 = max_row03(ws)  # количество строк

    i = 2  # мы точно знаем, что первая строка норм (чтобы не выйти за границы массива)
    while i < str_f1:
        if ws.cell(row=i, column=1).value.__str__().find('ТИПА') != -1:  # чтобы чекнуть строка 4561-4562
            # print(i)
            # print(ws.cell(row=i-1, column=2).value.__str__() + ws.cell(row=i, column=1).value.__str__())
            ws.cell(row=i - 1, column=2).value = ws.cell(row=i - 1, column=2).value.__str__() + \
                                                 ws.cell(row=i, column=1).value.__str__()
            ws.cell(row=i, column=1).value = None
        elif not is_number(ws.cell(row=i, column=1).value) and \
                not is_number_norm(ws['A' + i.__str__()].value.__str__().split(' ')[0].split('.')[0]):
            # если поднимаем больше чем на 1 строчку, то появляются None поэтому введем костыль
            crutch = 1
            for j in range(20):  # будем подниматься выше
                if is_number_norm(ws['A' + (i - j).__str__()].value.__str__().split(' ')[0].split('.')[0]):
                    break  # когда встретили изнамбернорм
                crutch += 1
            if ws.cell(row=i, column=1).value:
                ws.cell(row=i - crutch + 1, column=1).value = ws.cell(row=i - crutch + 1, column=1).value.__str__() + \
                                                              ws.cell(row=i, column=1).value.__str__()
                ws.cell(row=i, column=1).value = None
            for j in range(6, 11):  # поднимаем описание выше # чек 4623
                if ws.cell(row=i, column=j).value:  # если есть описание
                    ws.cell(row=i - crutch + 1, column=j).value = ws.cell(row=i, column=j).value.__str__()
                    ws.cell(row=i, column=j).value = None
        i += 1
    save_workbook(wb, FILE_NAME_OUT)
    # print('Таблица сохранена! 2из3')
    # end доочистка

    # start чистим пустые строки
    wbf1 = wb
    wsf1 = wbf1.active
    i = 1
    del_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    firstEmptyStr = 1
    while i < str_f1:
        if wsf1['A' + i.__str__()].value is None or wsf1['A' + i.__str__()].value == ' ':  # если
            # ячейка в А пустая - то удалить строку
            if firstEmptyStr == 1:
                firstEmptyStr = i
            # print('i - ' + i.__str__()) # строчки, которые нужно удалить
        if not (wsf1['A' + i.__str__()].value is None or wsf1['A' + i.__str__()].value == ' '):  # если
            # ячейка в А не пустая - то перенесем её
            j = i  # лишняя тема, но просто чтобы было читабельней (наверн)
            # print('j - ' + j.__str__()) # строчки, которые нужно переместить
            if firstEmptyStr != 1:
                for let in del_letters:
                    # wsf1[let + firstEmptyStr.__str__()].value = ''
                    wsf1[let + firstEmptyStr.__str__()].value = wsf1[let + j.__str__()].value
                    wsf1[let + j.__str__()].value = None
                firstEmptyStr += 1

        i += 1
    # end чистим пустые строки

    wbf1.save(FILE_NAME_OUT)  # сохраняем в новую табличку
    print('Таблица сохранена! 1/2')

    # start разделяем штуки на листы отдельные
    # FILE_NAME_IN = 'f1_edited.xlsx'
    # FILE_NAME_OUT = 'f1_edited_test.xlsx'

    wb = load_workbook(FILE_NAME_OUT)
    ws = wb.active

    list_count = 1  # найдем сначала количество листов, которое надо создать
    ws_temp = wb.create_sheet('Лист' + list_count.__str__())

    i = 1
    j = 1
    # print(ws.cell(row=1, column=1).value)
    # str_f1 = max_row03(ws)  # количество строк

    while i < str_f1:  # будем перебирать по каждой строке искать количество
        # нецифр (подряд идущие += 1) (((потом последний будем переносить в некст)))

        # print(i.__str__() +'---'+ is_number(ws.cell(row=i, column=1).value).__str__())
        if (is_number(ws.cell(row=i, column=1).value)) and (not is_number(ws.cell(row=i + 1, column=1).value)) and \
                (ws.cell(row=i + 1, column=6).value.__str__().split(' ')[0] == '1.' or
                 ws.cell(row=i + 2, column=6).value.__str__().split(' ')[0] == '1.' or
                 ws.cell(row=i + 3, column=6).value.__str__().split(' ')[0] == '1.'):  # если
            # это цифра а след не цифра, то создаем новый лист данные некст строки (включительно)
            # ((ласт - добавочное условие, если параметры меняются (Алексей попросил)))
            # переносим на новый лист
            j = 1
            list_count += 1
            ws_temp = wb.create_sheet('Лист' + list_count.__str__())
            # print(i.__str__() + '---' + ws.cell(row=i, column=1).value.__str__())
            # выше штука покажет где осуществляются разрывы на листы
            # ws = wbf1.create_sheet('Лист'+list_count.__str__())

        for col in range(1, 11):  # каждую строку переносим на текущий (temp) лист
            ws_temp.cell(row=j, column=col).value = ws.cell(row=i + 1, column=col).value  # мб j+1
        i += 1
        j += 1
    save_workbook(wb, FILE_NAME_OUT_LIST)
    print('Таблица сохранена! (ласт)')
    # end разделяем штуки на листы отдельные


def cleaner02(file_name):
    # редачить первый файл
    # FILE_NAME = 'Часть 02'

    FILE_NAME_IN = file_name + '.xlsx'
    FILE_NAME_OUT = file_name + ' исправленный.xlsx'

    wbf1 = load_workbook(FILE_NAME_IN)
    wsf1 = wbf1.active

    let_name = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    FJ_letters = ['C', 'F', 'G', 'H', 'I', 'J']
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    str_f1 = max_row(wsf1)  # количество строк
    # print('*************')

    # print(len(wsf1['B' + '4842'].value.__str__().split(' ')[0].split('.')[0]) <= 2)
    # print(len(wsf1['B' + '4843'].value.__str__().split(' ')[0].split('.')[0]) <= 2)
    # print(is_number_norm(wsf1['B' + '1'].value.__str__().split(' ')[0].split('.')[0]))
    # print(is_number_norm(wsf1['B' + '6'].value.__str__().split(' ')[0].split('.')[0]))

    i = 4
    while i < str_f1:
        wsf1['K' + i.__str__()].value = ''
        # print('i - ' + i.__str__())
        for j in range(i + 1, str_f1):
            # print('j - ' + j.__str__())
            # print(wsf1['A' + j.__str__()].value)
            # print(is_number(wsf1['A' + j.__str__()].value))
            if wsf1['A' + j.__str__()].value and is_number(wsf1[
                                                               'A' + j.__str__()].value) or j == str_f1 - 1:  # пос-
                # # леднее условие для того чтобы последнюю строку чистило
                # print('j - ' + j.__str__())
                if (j - i) > 1:
                    # print(i.__str__() + ' - ' + (j - i).__str__())  # строки с мусором

                    for k in range(i + 1, j):
                        # print(k)
                        # kostil = 0  # для того чтобы не двигал серию (Вроде хз)
                        # названия собираем и перемещаем в ячейку k (работает (Уже норм) криво)(Не везде
                        # работает: 37, 11354 - норм; 4, 42 - ненорм)
                        # is_number_norm(wsf1['B' + k.__str__()].value.__str__().split(' ')[0].split('.')[0]) and
                        if not is_number_norm(wsf1['B' + k.__str__()].value.__str__().split(' ')[0].split('.')[
                                                  0]):  # кароч, если не цифра. то значит не серия и часть названия,
                            # перемещаем её(часть) в сторону
                            if wsf1['B' + k.__str__()].value is not None:  # последнее условие чекате, чтобы
                                # "серия" не улетала с названием в право в ячеку К (мы сплитим по пробелам и по точкам
                                # если длина (наверн надо сделать Меньше 3 (типа вдруг двухциферная серия в начале)))
                                # print(is_number(wsf1['B' + k.__str__()].value.__str__().split(' ')[0].split('.')[0]))
                                # print(wsf1['B' + k.__str__()].value.__str__().split(' ')[0].split('.')[0])
                                for let in let_name:
                                    if wsf1[let + k.__str__()].value is not None:
                                        # print(let)
                                        wsf1['K' + i.__str__()].value = wsf1['K' + i.__str__()].value + wsf1[
                                            let + k.__str__()].value
                                        wsf1[let + k.__str__()].value = None

                        # Очистка столбца С (Завершено)
                        if (wsf1['B' + k.__str__()].value is None) and (wsf1['C' + k.__str__()].value is not None):
                            wsf1['C' + i.__str__()].value = wsf1['C' + i.__str__()].value + wsf1[
                                'C' + k.__str__()].value
                            wsf1['C' + k.__str__()].value = None

                        # Очистка столбца F-J (Завершено)
                        if wsf1['B' + k.__str__()].value is None:
                            for fj_let in FJ_letters:
                                if wsf1[fj_let + k.__str__()].value is not None and \
                                        wsf1[fj_let + i.__str__()].value is not None:
                                    wsf1[fj_let + i.__str__()].value = wsf1[fj_let + i.__str__()].value + wsf1[
                                        fj_let + k.__str__()].value
                                    wsf1[fj_let + k.__str__()].value = None

                        # for fj_let in FJ_letters:
                        #     if wsf1['B' + k.__str__()].value is None and wsf1[fj_let + k.__str__()].value is not None:
                        #         wsf1[fj_let + i.__str__()].value = wsf1[fj_let + i.__str__()].value + wsf1[
                        #             fj_let + k.__str__()].value
                        #         wsf1[fj_let + k.__str__()].value = None

                        # # настройки ниже нужны для особой записи какой-нибудь ячейки (нужно будет удалить букву из
                        # # списка FJ_letters)

                        # if wsf1['F' + k.__str__()].value is not None:
                        #     wsf1['F' + i.__str__()].value = wsf1['F' + i.__str__()].value + \
                        #                                     wsf1['F' + k.__str__()].value
                        #     wsf1['F' + k.__str__()].value = None
                        #
                        # # Очистка столбца G (Завершено)
                        # if wsf1['G' + k.__str__()].value is not None:
                        #     wsf1['G' + i.__str__()].value = wsf1['G' + i.__str__()].value + \
                        #                                     wsf1['G' + k.__str__()].value
                        #     wsf1['G' + k.__str__()].value = None
                        #
                        # # Очистка столбца H (Завершено)
                        # if wsf1['H' + k.__str__()].value is not None:
                        #     wsf1['H' + i.__str__()].value = wsf1['H' + i.__str__()].value + \
                        #                                     wsf1['H' + k.__str__()].value
                        #     wsf1['H' + k.__str__()].value = None
                        #
                        # # Очистка столбца I (Завершено)
                        # if wsf1['I' + k.__str__()].value is not None:
                        #     wsf1['I' + i.__str__()].value = wsf1['I' + i.__str__()].value + \
                        #                                     wsf1['I' + k.__str__()].value
                        #     wsf1['I' + k.__str__()].value = None
                        #
                        # # Очистка столбца J (Завершено) (чтобы Алексею было проще работать со
                        # # столбцом J)
                        # # будет добавить знак '|' или еще какой например в строке 11337 так будет удобнее)
                        # if wsf1['J' + k.__str__()].value is not None:
                        #     wsf1['J' + i.__str__()].value = wsf1['J' + i.__str__()].value + \
                        #                                     wsf1['J' + k.__str__()].value
                        #     wsf1['J' + k.__str__()].value = None

                # print('*************')
                i = j - 1
                break
        i += 1

    # print(wsf1['K' + '4841'].value)

    # print(wsf1['B' + '11340'].value)
    # print(wsf1['B' + '11346'].value)
    # print(wsf1['B' + '11293'].value)
    # wbf1.save("f2_edited.xlsx")  # сохраняем в новую табличку
    # print('Таблица сохранена!')
    #
    # # склеиваем и двигаем названия по типу "серия" влево (по образцу)
    # wbf1 = load_workbook('f2_edited.xlsx')
    # wsf1 = wbf1.active
    # str_f1 = max_row(wsf1, letters)  # количество строк
    i = 1
    while i < str_f1:
        if not (wsf1['A' + i.__str__()].value != ' ' and wsf1['A' + i.__str__()].value) and \
                wsf1['B' + i.__str__()].value:
            wsf1['A' + i.__str__()].value = ''
            for let in letters:
                if wsf1[let + i.__str__()].value:
                    # print(1)
                    wsf1['A' + i.__str__()].value = wsf1['A' + i.__str__()].value + wsf1[let + i.__str__()].value
                    wsf1[let + i.__str__()].value = None
        i += 1
    # wbf1.save("f2_edited_2.xlsx")  # сохраняем в новую табличку
    # print('Таблица сохранена!')

    # чистим пустые строки
    # wbf1 = load_workbook('f1_edited_2.xlsx')
    # wsf1 = wbf1.active

    # str_f1 = max_row(wsf1, letters)  # количество строк
    i = 1
    del_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    firstEmptyStr = 1
    while i < str_f1:
        if wsf1['A' + i.__str__()].value is None or \
                wsf1['A' + i.__str__()].value == ' ':  # если ячейка в А пустая - то удалить строку
            if firstEmptyStr == 1:
                firstEmptyStr = i
            print('i - ' + i.__str__())  # строчки, которые нужно удалить

        if not (wsf1['A' + i.__str__()].value is None or
                wsf1['A' + i.__str__()].value == ' '):  # если ячейка в А не пустая - то перенесем её
            j = i  # для читабельности
            print('j - ' + j.__str__())  # строчки, которые нужно переместить
            if firstEmptyStr != 1:
                for let in del_letters:
                    # wsf1[let + firstEmptyStr.__str__()].value = ''
                    wsf1[let + firstEmptyStr.__str__()].value = wsf1[let + j.__str__()].value
                    wsf1[let + j.__str__()].value = None
                firstEmptyStr += 1

        i += 1

    wbf1.save(FILE_NAME_OUT)  # сохраняем в новую табличку
    print('Таблица сохранена!')

    # wbf1.save("f1_edited_3.xlsx")  # сохраняем в новую табличку
    # print('Таблица сохранена!')

    #     в файле f1 ячейки E11271 и E11272 отредактированы вручную у f1
    #     заменить f1 на f2 и все будет работать (в начале(открытие) и в конце(сохранение))


def cleaner(file_name, type_of_cleaning):
    if type_of_cleaning == '02':
        cleaner02(file_name)
    elif type_of_cleaning == '03':
        cleaner03(file_name)


if __name__ == '__main__':

    # path = pathlib.Path('music.mp3')
    # print(path.exists())  # True
    # print(path.is_file())  # True

    FILE_NAME = ''
    TYPE_OF_CLEANING = ''

    check = False
    while not check:
        print('Введите название файла, который нужно очистить (без формата).')
        FILE_NAME = input()
        path = pathlib.Path(FILE_NAME + '.xlsx')
        if path.exists():
            check = True
        else:
            print('Необходимо ввести название файла, без формата (без .xlsx)')

    check2 = False
    while not check2:
        print('Какой тип очистки необходимо применить? Файл загрязнен по типу 02 или 03? [02/03]')
        TYPE_OF_CLEANING = input()
        if TYPE_OF_CLEANING == '02' or TYPE_OF_CLEANING == '03':
            check2 = True
        else:
            print('Вам необходимо ввести либо "02", либо "03"')

    cleaner(FILE_NAME, TYPE_OF_CLEANING)

# пример названий:

# Часть 05 Книга 1
# 03

# Часть 22 Книга 1
# 02
